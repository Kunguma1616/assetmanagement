from fastapi import APIRouter, HTTPException
import sys
import os
from datetime import datetime, timedelta
import openpyxl
from pathlib import Path
import pandas as pd

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from salesforce_service import SalesforceService
from lease_data_helper import get_lease_data_with_trade_groups

router = APIRouter(prefix="/api/cost", tags=["cost"])

# Get the path to Excel file
EXCEL_FILE = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), '..', 'HSBC_Leases.xlsx')

def load_lease_data():
    """Load lease data from Excel file"""
    if not os.path.exists(EXCEL_FILE):
        return []
    
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        ws = wb.active
        
        # Headers are in row 2
        headers = [ws.cell(2, i).value for i in range(1, ws.max_column + 1) if ws.cell(2, i).value]
        
        leases = []
        # Data starts from row 3
        for row_idx in range(3, ws.max_row + 1):
            row_data = {}
            for col_idx, header in enumerate(headers, 1):
                cell_value = ws.cell(row_idx, col_idx).value
                row_data[header] = cell_value
            
            # Only add if it has an identifier
            if row_data.get('Identifier ') or row_data.get('Registration Doc '):
                leases.append(row_data)
        
        return leases
    except Exception as e:
        print(f"❌ Error loading Excel file: {e}")
        return []


@router.get("/service/{van_number}")
def get_service_cost(van_number: str):
    """
    Get SERVICE COST & MAINTENANCE breakdown for a specific vehicle
    Returns: Van number, Trade Group, Total service cost, cost breakdown by type
    """
    try:
        sf = SalesforceService()
        
        print(f"\n🔧 FETCHING SERVICE & MAINTENANCE COSTS FOR VAN: {van_number}")
        print("="*80)
        
        # Find vehicle by van number with Trade Group
        vehicle_query = f"""
            SELECT Id, Name, Van_Number__c, Reg_No__c, Vehicle_Type__c, Status__c, Trade_Group__c
            FROM Vehicle__c
            WHERE Van_Number__c = '{van_number}'
            LIMIT 1
        """
        
        vehicles = sf.execute_soql(vehicle_query)
        
        if not vehicles:
            raise HTTPException(status_code=404, detail=f"Vehicle with van number {van_number} not found")
        
        vehicle = vehicles[0]
        vehicle_id = vehicle.get('Id')
        trade_group = vehicle.get('Trade_Group__c', 'N/A')
        
        print(f"✅ Found vehicle: {vehicle.get('Name')} ({van_number}) - Trade Group: {trade_group}")
        
        # Get detailed costs grouped by Type (ALL costs including Maintenance)
        cost_query = f"""
            SELECT Vehicle__c, Type__c, SUM(Payment_value__c) Total_By_Type
            FROM Vehicle_Service_Payment__c
            WHERE Vehicle__c = '{vehicle_id}'
            GROUP BY Vehicle__c, Type__c
        """
        
        print(f"🔍 Querying service & maintenance cost breakdown...")
        cost_records = sf.sf.query_all(cost_query).get('records', [])
        
        cost_by_type = {}
        total_cost = 0
        maintenance_cost = 0
        service_cost = 0
        
        # Parse breakdown
        for record in cost_records:
            cost_type = record.get('Type__c', 'Other')
            amount = float(record.get('Total_By_Type', 0) or 0)
            cost_by_type[cost_type] = round(amount, 2)
            total_cost += amount
            
            # Separate maintenance and service costs
            if cost_type == 'Maintenance':
                maintenance_cost += amount
            else:
                service_cost += amount
            
            print(f"  {cost_type}: £{amount:.2f}")
        
        monthly_avg = round(total_cost / 12, 2) if total_cost > 0 else 0
        
        print(f"\n💰 TOTAL COST: £{total_cost:.2f}")
        print(f"   Service Cost: £{service_cost:.2f}")
        print(f"   Maintenance Cost: £{maintenance_cost:.2f}")
        print(f"📊 Monthly Average: £{monthly_avg:.2f}")
        print("="*80 + "\n")
        
        return {
            "success": True,
            "vehicle": {
                "id": vehicle.get('Id'),
                "name": vehicle.get('Name'),
                "van_number": van_number,
                "registration": vehicle.get('Reg_No__c'),
                "vehicle_type": vehicle.get('Vehicle_Type__c'),
                "trade_group": trade_group,
                "status": vehicle.get('Status__c')
            },
            "costs": {
                "total_cost": round(total_cost, 2),
                "service_cost": round(service_cost, 2),
                "maintenance_cost": round(maintenance_cost, 2),
                "monthly_average": monthly_avg,
                "cost_by_type": cost_by_type,
                "cost_types_count": len(cost_by_type)
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Error fetching service costs: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/vehicle/{vehicle_id}")
def get_vehicle_cost(vehicle_id: str):
    """
    Get detailed cost data for a specific vehicle from Vehicle_Service_Payment__c
    Accepts: Vehicle ID (Salesforce ID) or Van Number
    """
    try:
        sf = SalesforceService()
        
        print(f"💰 Fetching cost data for vehicle: {vehicle_id}")
        
        # Try to find vehicle by ID or Van Number
        vehicle_query = f"""
            SELECT Id, Name, Van_Number__c, Reg_No__c, Vehicle_Type__c, Status__c
            FROM Vehicle__c
            WHERE Id = '{vehicle_id}' OR Van_Number__c = '{vehicle_id}'
            LIMIT 1
        """
        
        vehicles = sf.execute_soql(vehicle_query)
        
        if not vehicles:
            raise HTTPException(status_code=404, detail="Vehicle not found")
        
        vehicle = vehicles[0]
        vehicle_id_sf = vehicle.get('Id')
        
        print(f"✅ Found vehicle: {vehicle.get('Name')}")
        
        # Get all service payments for this vehicle from Vehicle_Service_Payment__c
        # This table has: Vehicle__c, Type__c (Insurance, MOT, Tax, Fuel, Maintenance, Repair, Rental, etc.), Payment_value__c
        cost_query = f"""
            SELECT Id, Type__c, Payment_value__c, CreatedDate, Description__c
            FROM Vehicle_Service_Payment__c
            WHERE Vehicle__c = '{vehicle_id_sf}'
            ORDER BY CreatedDate DESC
        """
        
        print(f"🔍 Querying service payments: {cost_query}")
        costs = sf.sf.query_all(cost_query).get('records', [])
        
        print(f"✅ Found {len(costs)} cost records")
        
        # Aggregate by type
        cost_by_type = {}
        total_cost = 0
        
        for cost in costs:
            cost_type = cost.get('Type__c', 'Other')
            value = float(cost.get('Payment_value__c', 0) or 0)
            
            if cost_type not in cost_by_type:
                cost_by_type[cost_type] = {'total': 0, 'count': 0, 'items': []}
            
            cost_by_type[cost_type]['total'] += value
            cost_by_type[cost_type]['count'] += 1
            cost_by_type[cost_type]['items'].append({
                'date': cost.get('CreatedDate', ''),
                'amount': value,
                'description': cost.get('Description__c', '')
            })
            
            total_cost += value
        
        print(f"💰 Total cost: £{total_cost}")
        print(f"📊 Cost breakdown: {list(cost_by_type.keys())}")
        
        # Calculate specific cost types
        maintenance_cost = cost_by_type.get('Maintenance', {}).get('total', 0)
        repair_cost = cost_by_type.get('Repair', {}).get('total', 0)
        fuel_cost = cost_by_type.get('Fuel', {}).get('total', 0)
        insurance_cost = cost_by_type.get('Insurance', {}).get('total', 0)
        mot_cost = cost_by_type.get('MOT', {}).get('total', 0)
        tax_cost = cost_by_type.get('Tax', {}).get('total', 0)
        rental_cost = cost_by_type.get('Rental', {}).get('total', 0)
        
        monthly_avg = round(total_cost / 12, 2) if total_cost > 0 else 0
        
        return {
            "success": True,
            "vehicle": {
                "id": vehicle.get('Id'),
                "name": vehicle.get('Name'),
                "van_number": vehicle.get('Van_Number__c'),
                "reg_no": vehicle.get('Reg_No__c'),
                "type": vehicle.get('Vehicle_Type__c'),
                "status": vehicle.get('Status__c')
            },
            "costs": {
                "total_cost": round(total_cost, 2),
                "maintenance_cost": round(maintenance_cost, 2),
                "repair_cost": round(repair_cost, 2),
                "fuel_cost": round(fuel_cost, 2),
                "insurance_cost": round(insurance_cost, 2),
                "mot_cost": round(mot_cost, 2),
                "tax_cost": round(tax_cost, 2),
                "rental_cost": round(rental_cost, 2),
                "monthly_average": monthly_avg,
                "cost_by_type": {k: round(v['total'], 2) for k, v in cost_by_type.items()},
                "cost_count": len(costs)
            },
            "detailed_costs": cost_by_type,
            "cost_items": costs
        }
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Error fetching vehicle cost: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/all-vehicles")
def get_all_vehicles_costs():
    """
    Get total lifecycle cost and breakdown for ALL vehicles
    Shows: Total cost per vehicle, cost by type (Insurance, MOT, Rental, etc.)
    """
    try:
        sf = SalesforceService()
        
        print("\n" + "="*80)
        print("💰 FETCHING ALL VEHICLE COSTS")
        print("="*80)
        
        # Step 1: Get all vehicles
        vehicle_query = """
            SELECT Id, Name, Van_Number__c, Reg_No__c, Vehicle_Type__c, Status__c
            FROM Vehicle__c
            ORDER BY Van_Number__c ASC
        """
        
        vehicles = sf.sf.query_all(vehicle_query).get('records', [])
        print(f"✅ Found {len(vehicles)} vehicles")
        
        # Step 2: Get ALL service payments (don't use GROUP BY - fetch all and aggregate in Python)
        cost_query = """
            SELECT Vehicle__c, Type__c, Payment_value__c, CreatedDate
            FROM Vehicle_Service_Payment__c
            ORDER BY Vehicle__c, Type__c
        """
        
        print(f"🔍 Querying service payments...")
        cost_records = sf.sf.query_all(cost_query).get('records', [])
        print(f"✅ Found {len(cost_records)} cost records")
        
        # Aggregate costs by vehicle and type in Python
        vehicle_costs = {}
        for cost in cost_records:
            vehicle_id = cost.get('Vehicle__c')
            cost_type = cost.get('Type__c', 'Other')
            amount = float(cost.get('Payment_value__c', 0) or 0)
            
            if vehicle_id not in vehicle_costs:
                vehicle_costs[vehicle_id] = {
                    'total': 0,
                    'by_type': {}
                }
            
            vehicle_costs[vehicle_id]['total'] += amount
            if cost_type not in vehicle_costs[vehicle_id]['by_type']:
                vehicle_costs[vehicle_id]['by_type'][cost_type] = 0
            vehicle_costs[vehicle_id]['by_type'][cost_type] += amount
        
        print(f"📊 Processed costs for {len(vehicle_costs)} vehicles with costs")
        
        # Step 3: Build response with all vehicles and their costs
        vehicles_list = []
        total_fleet_cost = 0
        fleet_cost_by_type = {}
        
        for vehicle in vehicles:
            vehicle_id = vehicle.get('Id')
            van_number = vehicle.get('Van_Number__c')
            
            # Get costs for this vehicle
            vehicle_cost_info = vehicle_costs.get(vehicle_id, {'total': 0, 'by_type': {}})
            total_cost = vehicle_cost_info['total']
            cost_breakdown = vehicle_cost_info['by_type']
            
            total_fleet_cost += total_cost
            
            # Aggregate fleet costs by type
            for cost_type, amount in cost_breakdown.items():
                if cost_type not in fleet_cost_by_type:
                    fleet_cost_by_type[cost_type] = 0
                fleet_cost_by_type[cost_type] += amount
            
            vehicles_list.append({
                "vehicle_id": vehicle_id,
                "name": vehicle.get('Name'),
                "van_number": van_number,
                "registration": vehicle.get('Reg_No__c'),
                "vehicle_type": vehicle.get('Vehicle_Type__c'),
                "status": vehicle.get('Status__c'),
                "total_cost": round(total_cost, 2),
                "cost_breakdown": {k: round(v, 2) for k, v in cost_breakdown.items()},
                "monthly_average": round(total_cost / 12, 2) if total_cost > 0 else 0
            })
        
        # Sort by total cost (highest first)
        vehicles_list.sort(key=lambda x: x['total_cost'], reverse=True)
        
        avg_vehicle_cost = round(total_fleet_cost / len(vehicles), 2) if vehicles else 0
        
        print(f"\n" + "="*80)
        print(f"💰 FLEET COST SUMMARY")
        print(f"   Total Fleet Cost: £{total_fleet_cost:.2f}")
        print(f"   Average per Vehicle: £{avg_vehicle_cost:.2f}")
        print(f"   Total Vehicles: {len(vehicles)}")
        print(f"   Vehicles with Costs: {len(vehicle_costs)}")
        print("="*80 + "\n")
        
        return {
            "success": True,
            "summary": {
                "total_fleet_cost": round(total_fleet_cost, 2),
                "average_vehicle_cost": avg_vehicle_cost,
                "vehicle_count": len(vehicles),
                "vehicles_with_costs": len(vehicle_costs)
            },
            "cost_breakdown_by_type": {k: round(v, 2) for k, v in fleet_cost_by_type.items()},
            "vehicles": vehicles_list,
            "top_cost_vehicles": vehicles_list[:10]
        }
        
    except Exception as e:
        print(f"❌ Error fetching vehicle costs: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/service-maintenance/insights")
def get_service_maintenance_insights(trade_filter: str = None):
    """
    Get SERVICE & MAINTENANCE cost insights for all vehicles
    Shows: Van number, Trade Group, Total service cost, maintenance cost, breakdown
    Includes team member information and highest cost vehicles
    For: Stakeholders & Fleet Management Team
    
    Optional: trade_filter - If provided, only returns data for that trade group
    """
    try:
        sf = SalesforceService()
        
        print("\n" + "="*80)
        print("💰 FETCHING SERVICE & MAINTENANCE COST INSIGHTS")
        if trade_filter:
            print(f"🔒 RESTRICTED VIEW: {trade_filter} only")
        print("="*80)
        
        # Team member information (removed Peter - no longer needed)
        team_members = [
            {
                "name": "Martin Mackie",
                "email": "martin.mackie@aspect.co.uk",
                "trade": "Drainage & Plumbing",
                "role": "Fleet Manager"
            },
            {
                "name": "James Parkinson",
                "email": "james.parkinson@aspect.co.uk",
                "trade": "HVAC & Electrical",
                "role": "Operations Lead"
            },
            {
                "name": "Lee Merryweather",
                "email": "lee.merryweather@aspect.co.uk",
                "trade": "Building Fabric",
                "role": "Maintenance Coordinator"
            },
            {
                "name": "Marjan",
                "email": "marjan@aspect.co.uk",
                "trade": "LDR",
                "role": "Cost Analyst"
            }
        ]
        
        # Step 1: Get all vehicles with Trade Group
        vehicle_query = """
            SELECT Id, Name, Van_Number__c, Reg_No__c, Vehicle_Type__c, Status__c, Trade_Group__c
            FROM Vehicle__c
            ORDER BY Van_Number__c ASC
        """
        
        vehicles = sf.sf.query_all(vehicle_query).get('records', [])
        print(f"✅ Found {len(vehicles)} vehicles")
        
        # Step 2: Get ALL service & maintenance payments with Trade Group rollup
        cost_query = """
            SELECT Vehicle__c, Type__c, SUM(Payment_value__c) Total_Amount
            FROM Vehicle_Service_Payment__c
            GROUP BY Vehicle__c, Type__c
            ORDER BY Vehicle__c, Type__c
        """
        
        print(f"🔍 Querying service & maintenance costs...")
        cost_records = sf.sf.query_all(cost_query).get('records', [])
        print(f"✅ Found {len(cost_records)} cost records")
        
        # Aggregate by vehicle
        vehicle_costs_map = {}
        for cost in cost_records:
            vehicle_id = cost.get('Vehicle__c')
            cost_type = cost.get('Type__c', 'Other')
            amount = float(cost.get('Total_Amount', 0) or 0)
            
            if vehicle_id not in vehicle_costs_map:
                vehicle_costs_map[vehicle_id] = {
                    'total_cost': 0,
                    'service_cost': 0,
                    'maintenance_cost': 0,
                    'by_type': {}
                }
            
            vehicle_costs_map[vehicle_id]['by_type'][cost_type] = round(amount, 2)
            vehicle_costs_map[vehicle_id]['total_cost'] += amount
            
            if cost_type == 'Maintenance':
                vehicle_costs_map[vehicle_id]['maintenance_cost'] += amount
            else:
                vehicle_costs_map[vehicle_id]['service_cost'] += amount
        
        print(f"📊 Processed costs for {len(vehicle_costs_map)} vehicles")
        
        # Step 3: Build response with all vehicles and insights
        vehicles_list = []
        total_fleet_cost = 0
        total_maintenance = 0
        total_service = 0
        
        for vehicle in vehicles:
            vehicle_id = vehicle.get('Id')
            van_number = vehicle.get('Van_Number__c') or 'N/A'
            trade_group = vehicle.get('Trade_Group__c') or 'Not Assigned'
            
            # If trade_filter is specified, skip vehicles not in that trade group
            if trade_filter and trade_group != trade_filter:
                continue
            
            # Get costs for this vehicle
            if vehicle_id in vehicle_costs_map:
                costs = vehicle_costs_map[vehicle_id]
                total_cost = costs['total_cost']
                maintenance_cost = costs['maintenance_cost']
                service_cost = costs['service_cost']
                cost_breakdown = costs['by_type']
            else:
                total_cost = 0
                maintenance_cost = 0
                service_cost = 0
                cost_breakdown = {}
            
            total_fleet_cost += total_cost
            total_maintenance += maintenance_cost
            total_service += service_cost
            
            vehicles_list.append({
                "vehicle_id": vehicle_id,
                "name": vehicle.get('Name'),
                "van_number": van_number,
                "registration": vehicle.get('Reg_No__c'),
                "vehicle_type": vehicle.get('Vehicle_Type__c'),
                "trade_group": trade_group,
                "status": vehicle.get('Status__c'),
                "costs": {
                    "total_cost": round(total_cost, 2),
                    "service_cost": round(service_cost, 2),
                    "maintenance_cost": round(maintenance_cost, 2),
                    "monthly_average": round(total_cost / 12, 2) if total_cost > 0 else 0,
                    "cost_by_type": cost_breakdown
                }
            })
        
        # Sort by maintenance cost (highest first)
        vehicles_list.sort(key=lambda x: x['costs']['maintenance_cost'], reverse=True)
        
        # Get top 5 vehicles with highest maintenance costs
        top_maintenance_vehicles = vehicles_list[:5]
        
        # Get top 5 vehicles with highest service costs
        top_service_vehicles = sorted(vehicles_list, key=lambda x: x['costs']['service_cost'], reverse=True)[:5]
        
        avg_vehicle_cost = round(total_fleet_cost / len(vehicles), 2) if vehicles else 0
        
        print(f"\n" + "="*80)
        print(f"💰 SERVICE & MAINTENANCE COST SUMMARY")
        print(f"   Total Fleet Cost: £{total_fleet_cost:.2f}")
        print(f"   Service Cost: £{total_service:.2f}")
        print(f"   Maintenance Cost: £{total_maintenance:.2f}")
        print(f"   Average per Vehicle: £{avg_vehicle_cost:.2f}")
        print(f"   Total Vehicles: {len(vehicles)}")
        print(f"   Vehicles with Costs: {len(vehicle_costs_map)}")
        print("="*80 + "\n")
        
        return {
            "success": True,
            "summary": {
                "total_fleet_cost": round(total_fleet_cost, 2),
                "total_service_cost": round(total_service, 2),
                "total_maintenance_cost": round(total_maintenance, 2),
                "average_vehicle_cost": round(total_fleet_cost / len(vehicles_list), 2) if vehicles_list else 0,
                "vehicle_count": len(vehicles_list),
                "vehicles_with_costs": len([v for v in vehicles_list if v['costs']['total_cost'] > 0])
            },
            "team_members": team_members,
            "access_control": {
                "is_restricted": trade_filter is not None,
                "restricted_to_trade": trade_filter,
                "visible_vehicles": len(vehicles_list)
            },
            "insights": {
                "top_maintenance_vehicles": sorted(vehicles_list, key=lambda x: x['costs']['maintenance_cost'], reverse=True)[:5],
                "top_service_vehicles": sorted(vehicles_list, key=lambda x: x['costs']['service_cost'], reverse=True)[:5],
                "highest_maintenance_cost": max([v['costs']['maintenance_cost'] for v in vehicles_list], default=0),
                "highest_service_cost": max([v['costs']['service_cost'] for v in vehicles_list], default=0),
            },
            "vehicles": vehicles_list,
            "timestamp": datetime.now().isoformat()
        }
        
    except Exception as e:
        print(f"❌ Error fetching insights: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/leases/all")
def get_all_leases():
    """Get all lease data from Excel file"""
    try:
        leases = load_lease_data()
        
        if not leases:
            return {
                "success": False,
                "total": 0,
                "leases": []
            }
        
        # Calculate totals
        total_vehicles = len(leases)
        total_capital_cost = sum(float(lease.get('Capital Cost ', 0) or 0) for lease in leases)
        
        print(f"\n✅ LOADED {total_vehicles} LEASES FROM EXCEL")
        print(f"   Total Capital Cost: £{total_capital_cost:.2f}")
        
        return {
            "success": True,
            "total": total_vehicles,
            "total_capital_cost": round(total_capital_cost, 2),
            "average_capital_cost": round(total_capital_cost / total_vehicles, 2) if total_vehicles > 0 else 0,
            "leases": [
                {
                    "identifier": lease.get('Identifier ', '').strip() if lease.get('Identifier ') else None,
                    "vehicle_type": lease.get('Type', ''),
                    "contract_number": lease.get('Contract number'),
                    "registration": lease.get('Registration Doc ', '').strip() if lease.get('Registration Doc ') else None,
                    "make_model": lease.get('Make and Model ', '').strip() if lease.get('Make and Model ') else None,
                    "start_date": lease.get('Agreement Start Date ', None),
                    "end_date": lease.get('Agreement end date ', None),
                    "term_months": lease.get('Agreement term (months)', None),
                    "net_capital": float(lease.get('Net Capital ', 0) or 0),
                    "vat": float(lease.get('VAT on Acquisition ', 0) or 0),
                    "capital_cost": float(lease.get('Capital Cost ', 0) or 0),
                    "monthly_payment": round(float(lease.get('Capital Cost ', 0) or 0) / (lease.get('Agreement term (months)', 1) or 1), 2)
                }
                for lease in leases
            ]
        }
    except Exception as e:
        print(f"❌ Error loading leases: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/leases/with-trade-group")
def get_leases_with_trade_group(trade_filter: str = None):
    """Get HSBC lease data with trade group information and optional filtering"""
    try:
        leases = load_lease_data()
        
        if not leases:
            return {
                "success": False,
                "total": 0,
                "leases": []
            }
        
        # Create a mapping of registration to Salesforce vehicles for trade group lookup
        sf = SalesforceService()
        vehicle_query = """
            SELECT Reg_No__c, Trade_Group__c
            FROM Vehicle__c
            WHERE Reg_No__c != NULL
        """
        
        vehicles = sf.sf.query_all(vehicle_query).get('records', [])
        reg_to_trade = {v.get('Reg_No__c'): v.get('Trade_Group__c', 'Not Assigned') for v in vehicles}
        
        # Process leases
        lease_list = []
        total_capital_cost = 0
        
        for lease in leases:
            reg = (lease.get('Registration Doc ', '') or '').strip().upper()
            trade_group = reg_to_trade.get(reg, 'Not Assigned')
            
            # Apply trade filter if specified
            if trade_filter and trade_group != trade_filter:
                continue
            
            capital_cost = float(lease.get('Capital Cost ', 0) or 0)
            term_months = lease.get('Agreement term (months)', 1) or 1
            
            lease_list.append({
                "identifier": lease.get('Identifier ', '').strip() if lease.get('Identifier ') else None,
                "vehicle_type": lease.get('Type', ''),
                "contract_number": lease.get('Contract number'),
                "registration": reg if reg else None,
                "make_model": lease.get('Make and Model ', '').strip() if lease.get('Make and Model ') else None,
                "trade_group": trade_group,
                "start_date": lease.get('Agreement Start Date ', None),
                "end_date": lease.get('Agreement end date ', None),
                "term_months": term_months,
                "net_capital": float(lease.get('Net Capital ', 0) or 0),
                "vat": float(lease.get('VAT on Acquisition ', 0) or 0),
                "capital_cost": round(capital_cost, 2),
                "monthly_payment": round(capital_cost / term_months, 2) if term_months > 0 else 0
            })
            
            total_capital_cost += capital_cost
        
        print(f"\n✅ LOADED {len(lease_list)} HSBC LEASES")
        if trade_filter:
            print(f"   Filtered to trade: {trade_filter}")
        print(f"   Total Capital Cost: £{total_capital_cost:.2f}")
        
        return {
            "success": True,
            "total": len(lease_list),
            "total_capital_cost": round(total_capital_cost, 2),
            "average_capital_cost": round(total_capital_cost / len(lease_list), 2) if lease_list else 0,
            "leases": lease_list
        }
    except Exception as e:
        print(f"❌ Error loading leases: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/leases/registration/{registration}")
def get_lease_by_registration(registration: str):
    """Get lease data for a specific vehicle registration"""
    try:
        leases = load_lease_data()
        
        # Search for matching registration
        matching_lease = None
        for lease in leases:
            reg = (lease.get('Registration Doc ', '') or '').strip().upper()
            if reg == registration.upper():
                matching_lease = lease
                break
        
        if not matching_lease:
            return {
                "success": False,
                "message": f"No lease found for registration {registration}"
            }
        
        capital_cost = float(matching_lease.get('Capital Cost ', 0) or 0)
        term_months = matching_lease.get('Agreement term (months)', 1) or 1
        
        return {
            "success": True,
            "identifier": matching_lease.get('Identifier ', '').strip() if matching_lease.get('Identifier ') else None,
            "vehicle_type": matching_lease.get('Type', ''),
            "contract_number": matching_lease.get('Contract number'),
            "registration": matching_lease.get('Registration Doc ', '').strip() if matching_lease.get('Registration Doc ') else None,
            "make_model": matching_lease.get('Make and Model ', '').strip() if matching_lease.get('Make and Model ') else None,
            "start_date": matching_lease.get('Agreement Start Date ', None),
            "end_date": matching_lease.get('Agreement end date ', None),
            "term_months": term_months,
            "net_capital": float(matching_lease.get('Net Capital ', 0) or 0),
            "vat": float(matching_lease.get('VAT on Acquisition ', 0) or 0),
            "capital_cost": round(capital_cost, 2),
            "monthly_payment": round(capital_cost / term_months, 2)
        }
    except Exception as e:
        print(f"❌ Error fetching lease: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/leases/csv-all")
def get_all_csv_leases():
    """Read ALL rows from the HSBC Excel file and return as JSON (no filtering)"""
    EXCEL_PATHS = [
        Path(r"C:\Users\Kunguma.Balaji\Downloads\HSBC_Leases_Fixed.xlsx"),
        Path(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))) / '..' / 'HSBC_Leases_Fixed.xlsx',
        Path(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))) / '..' / 'HSBC_Leases.xlsx',
    ]

    CURRENCY_COLS = [
        'Net Capital', 'VAT on Acquisition', 'RFL', 'Capital Cost',
        'Arrangement Fee', 'Finance Interest', 'Initial Payment',
        'Monthly Installment', 'Final Payment', 'Total Repayment',
    ]

    def fmt_currency(val):
        try:
            if pd.isna(val):
                return ''
            f = float(val)
            return '' if f == 0 else f'£{f:,.2f}'
        except Exception:
            v = str(val).strip() if val is not None else ''
            return v if v and v not in ('0', '0.0', '-') else ''

    try:
        excel_file = None
        for p in EXCEL_PATHS:
            if p.exists():
                excel_file = str(p.resolve())
                break

        if not excel_file:
            return {"success": False, "total": 0, "rows": [],
                    "error": f"Excel file not found. Searched: {[str(p) for p in EXCEL_PATHS]}"}

        sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        from excel_handler import read_and_clean_hsbc_leases
        df = read_and_clean_hsbc_leases(excel_file, verbose=False)

        # Forward-fill Identifier for merged/blank rows
        if 'Identifier' in df.columns:
            df['Identifier'] = df['Identifier'].ffill()

        rows = []
        for _, row_data in df.iterrows():
            r = {}
            for col in df.columns:
                val = row_data[col]
                if col in CURRENCY_COLS:
                    r[col] = fmt_currency(val)
                elif hasattr(val, 'strftime'):
                    try:
                        r[col] = val.strftime('%d/%m/%Y')
                    except Exception:
                        r[col] = ''
                else:
                    try:
                        r[col] = '' if pd.isna(val) else (str(val) if not isinstance(val, str) else val)
                    except Exception:
                        r[col] = str(val) if val is not None else ''

            # Skip rows with no Type and no Make and Model
            if not r.get('Type') and not r.get('Make and Model'):
                continue

            r['_identifier'] = r.get('Identifier', '')
            rows.append(r)

        print(f"✅ Excel leases loaded: {len(rows)} rows from {excel_file}")
        return {"success": True, "total": len(rows), "rows": rows}

    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


# ─── Fleet Operational Insights ───────────────────────────────────────────────
from operational_insights import get_operational_insights, get_top_10_expensive_vans, get_cost_summary


@router.get("/operational-insights")
def get_fleet_insights(trade_group: str = None):
    """
    Get combined operational costs (Lease + Service + Maintenance) for all vehicles
    Returns all vehicles sorted by total operational cost (highest first)
    Supports optional trade_group filter
    """
    try:
        insights = get_operational_insights(trade_group)
        summary = get_cost_summary(trade_group)
        
        return {
            "success": True,
            "summary": summary,
            "vehicles": insights,
            "total": len(insights)
        }
    except Exception as e:
        print(f"❌ Error getting operational insights: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/operational-insights/top-10")
def get_top_10_vans(trade_group: str = None):
    """
    Get top 10 most expensive vans to operate
    Supports optional trade_group filter
    """
    try:
        top_10 = get_top_10_expensive_vans(trade_group)
        summary = get_cost_summary(trade_group)
        
        return {
            "success": True,
            "summary": summary,
            "vehicles": top_10,
            "total": len(top_10)
        }
    except Exception as e:
        print(f"❌ Error getting top 10 vans: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/operational-insights/cost-breakdown")
def get_cost_breakdown(trade_group: str = None):
    """
    Get cost breakdown (Lease vs Service vs Maintenance) for visualization
    """
    try:
        insights = get_operational_insights(trade_group)
        
        # Calculate breakdown percentages
        total_lease = sum(v['lease_cost'] for v in insights)
        total_service = sum(v['service_cost'] for v in insights)
        total_maintenance = sum(v['maintenance_cost'] for v in insights)
        total = total_lease + total_service + total_maintenance
        
        breakdown = {
            'Lease Cost': round(total_lease, 2) if total_lease > 0 else 0,
            'Service Cost': round(total_service, 2) if total_service > 0 else 0,
            'Maintenance Cost': round(total_maintenance, 2) if total_maintenance > 0 else 0,
        }
        
        return {
            "success": True,
            "breakdown": breakdown,
            "total": round(total, 2),
            "percentages": {
                'Lease Cost': round((total_lease / total * 100), 1) if total > 0 else 0,
                'Service Cost': round((total_service / total * 100), 1) if total > 0 else 0,
                'Maintenance Cost': round((total_maintenance / total * 100), 1) if total > 0 else 0,
            }
        }
    except Exception as e:
        print(f"❌ Error getting cost breakdown: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ════════════════════════════════════════════════════════════════════════════════
# NEW ENDPOINT: COMPREHENSIVE VEHICLE FINANCIAL OVERVIEW
# Combines: Lease Capital Cost + Service Costs + Maintenance Costs
# ════════════════════════════════════════════════════════════════════════════════

@router.get("/vehicle-financial-overview")
def get_vehicle_financial_overview(trade_group: str = None):
    """
    Get comprehensive cost breakdown for all vehicles showing:
    - Lease Capital Cost (from HSBC leases)
    - Total Service Cost
    - Total Maintenance Cost
    - Total Cost Breakdown
    - Where the money went (percentages)
    
    This answers: For each van, what did we spend on capital vs operations?
    
    Args:
        trade_group: Optional filter by trade group
    
    Returns:
        Comprehensive financial overview per vehicle sorted by total cost
    """
    try:
        print("\n" + "="*80)
        print("📊 GENERATING COMPREHENSIVE VEHICLE FINANCIAL OVERVIEW")
        print("="*80)
        
        sf = SalesforceService()
        
        # Step 1: Load lease data indexed by registration number
        print("\n📋 Loading lease data...")
        lease_df = get_lease_data_with_trade_groups(as_dict=False)
        lease_dict = {}
        
        if lease_df is not None and len(lease_df) > 0:
            for _, row in lease_df.iterrows():
                reg_key = str(row.get('Registration Doc', '')).strip().upper() if pd.notna(row.get('Registration Doc')) else None
                if reg_key:
                    lease_dict[reg_key] = {
                        'identifier': row.get('Identifier'),
                        'capital_cost': float(row.get('Capital Cost', 0)) if pd.notna(row.get('Capital Cost')) else 0,
                        'net_capital': float(row.get('Net Capital', 0)) if pd.notna(row.get('Net Capital')) else 0,
                        'repayment': float(row.get('Total Repayment', 0)) if pd.notna(row.get('Total Repayment')) else 0,
                        'asset_type': row.get('Asset Type', 'Unknown')
                    }
            print(f"✅ Loaded {len(lease_dict)} lease records")
        else:
            print("⚠️  No lease data available")
        
        # Step 2: Get all vehicles from Salesforce
        print("\n🚗 Fetching vehicles from Salesforce...")
        vehicle_query = """
            SELECT Id, Name, Van_Number__c, Reg_No__c, Vehicle_Type__c, Trade_Group__c, Status__c
            FROM Vehicle__c
            WHERE Status__c = 'Active'
            ORDER BY Van_Number__c ASC
        """
        
        vehicles = sf.execute_soql(vehicle_query)
        print(f"✅ Found {len(vehicles)} active vehicles")
        
        # Step 3: Get all service & maintenance costs
        print("\n💰 Fetching service & maintenance costs...")
        cost_query = """
            SELECT Vehicle__c, Type__c, SUM(Payment_value__c) Total_By_Type
            FROM Vehicle_Service_Payment__c
            GROUP BY Vehicle__c, Type__c
        """
        
        cost_records = sf.sf.query_all(cost_query).get('records', [])
        
        # Map costs by vehicle
        vehicle_costs_map = {}
        for record in cost_records:
            vehicle_id = record.get('Vehicle__c')
            cost_type = record.get('Type__c', 'Other')
            amount = float(record.get('Total_By_Type', 0) or 0)
            
            if vehicle_id not in vehicle_costs_map:
                vehicle_costs_map[vehicle_id] = {
                    'service_cost': 0,
                    'maintenance_cost': 0,
                    'total_ops_cost': 0,
                    'cost_breakdown': {}
                }
            
            vehicle_costs_map[vehicle_id]['cost_breakdown'][cost_type] = round(amount, 2)
            vehicle_costs_map[vehicle_id]['total_ops_cost'] += amount
            
            if cost_type == 'Maintenance':
                vehicle_costs_map[vehicle_id]['maintenance_cost'] += amount
            else:
                vehicle_costs_map[vehicle_id]['service_cost'] += amount
        
        print(f"✅ Retrieved costs for {len(vehicle_costs_map)} vehicles")
        
        # Step 4: Build comprehensive vehicle financial records
        print("\n🔧 Compiling financial overview...")
        vehicles_financial = []
        total_fleet_capital = 0
        total_fleet_operations = 0
        
        for vehicle in vehicles:
            vehicle_id = vehicle.get('Id')
            van_number = vehicle.get('Van_Number__c') or 'UNASSIGNED'
            reg_number = (vehicle.get('Reg_No__c') or '').upper()
            trade_group = vehicle.get('Trade_Group__c') or 'Not Assigned'
            vehicle_name = vehicle.get('Name', f'Van {van_number}')
            
            # Apply trade group filter if provided
            if trade_group and trade_group.lower() != 'not assigned':
                if trade_group and trade_group.lower() != 'not assigned':
                    if trade_group and trade_group.lower() != 'not assigned':
                        pass  # Keep vehicle
            
            if trade_group and trade_group.lower() != 'not assigned':
                pass  # Will filter below
            
            # Get lease data
            capital_cost = 0
            net_capital = 0
            repayment = 0
            asset_type = 'Unknown'
            identifier = None
            
            if reg_number and reg_number in lease_dict:
                lease_info = lease_dict[reg_number]
                capital_cost = lease_info['capital_cost']
                net_capital = lease_info['net_capital']
                repayment = lease_info['repayment']
                asset_type = lease_info['asset_type']
                identifier = lease_info['identifier']
            
            # Get operational costs
            service_cost = 0
            maintenance_cost = 0
            total_ops_cost = 0
            cost_breakdown = {}
            
            if vehicle_id in vehicle_costs_map:
                costs = vehicle_costs_map[vehicle_id]
                service_cost = costs['service_cost']
                maintenance_cost = costs['maintenance_cost']
                total_ops_cost = costs['total_ops_cost']
                cost_breakdown = costs['cost_breakdown']
            
            # Calculate totals
            total_cost = capital_cost + total_ops_cost
            
            # Build vehicle record
            vehicle_financial = {
                'van_number': van_number,
                'registration': reg_number or 'N/A',
                'vehicle_name': vehicle_name,
                'trade_group': trade_group,
                'identifier': identifier or 'N/A',
                'asset_type': asset_type,
                'capital_cost': round(capital_cost, 2),
                'net_capital': round(net_capital, 2),
                'lease_repayment': round(repayment, 2),
                'service_cost': round(service_cost, 2),
                'maintenance_cost': round(maintenance_cost, 2),
                'total_operations_cost': round(total_ops_cost, 2),
                'total_cost': round(total_cost, 2),
                'cost_breakdown': {k: round(v, 2) for k, v in cost_breakdown.items()},
                'cost_distribution': {
                    'capital': round(capital_cost, 2) if total_cost > 0 else 0,
                    'operations': round(total_ops_cost, 2) if total_cost > 0 else 0,
                }
            }
            
            # Cost percentages
            if total_cost > 0:
                vehicle_financial['cost_percentage'] = {
                    'capital': round((capital_cost / total_cost * 100), 1),
                    'service': round((service_cost / total_cost * 100), 1),
                    'maintenance': round((maintenance_cost / total_cost * 100), 1)
                }
            else:
                vehicle_financial['cost_percentage'] = {'capital': 0, 'service': 0, 'maintenance': 0}
            
            # Only include if has costs or lease data
            if total_cost > 0 or capital_cost > 0:
                vehicles_financial.append(vehicle_financial)
                total_fleet_capital += capital_cost
                total_fleet_operations += total_ops_cost
        
        # Apply trade group filter if specified
        if trade_group:
            vehicles_financial = [v for v in vehicles_financial if v['trade_group'].lower() == trade_group.lower()]
        
        # Sort by total cost (highest first)
        vehicles_financial.sort(key=lambda x: x['total_cost'], reverse=True)
        
        # Calculate fleet summary
        total_fleet_cost = total_fleet_capital + total_fleet_operations
        
        summary = {
            'total_vehicles_with_costs': len(vehicles_financial),
            'total_fleet_capital_cost': round(total_fleet_capital, 2),
            'total_fleet_operations_cost': round(total_fleet_operations, 2),
            'total_fleet_service_cost': round(sum(v['service_cost'] for v in vehicles_financial), 2),
            'total_fleet_maintenance_cost': round(sum(v['maintenance_cost'] for v in vehicles_financial), 2),
            'total_fleet_cost': round(total_fleet_cost, 2),
            'average_capital_per_vehicle': round(total_fleet_capital / len(vehicles_financial), 2) if vehicles_financial else 0,
            'average_operations_per_vehicle': round(total_fleet_operations / len(vehicles_financial), 2) if vehicles_financial else 0,
            'fleet_cost_distribution': {
                'capital': round(total_fleet_capital, 2),
                'operations': round(total_fleet_operations, 2)
            }
        }
        
        # Fleet-wide percentages
        if total_fleet_cost > 0:
            summary['fleet_cost_percentage'] = {
                'capital': round((total_fleet_capital / total_fleet_cost * 100), 1),
                'operations': round((total_fleet_operations / total_fleet_cost * 100), 1),
                'service': round((summary['total_fleet_service_cost'] / total_fleet_cost * 100), 1),
                'maintenance': round((summary['total_fleet_maintenance_cost'] / total_fleet_cost * 100), 1)
            }
        else:
            summary['fleet_cost_percentage'] = {'capital': 0, 'operations': 0, 'service': 0, 'maintenance': 0}
        
        print(f"\n✅ FINANCIAL OVERVIEW SUMMARY")
        print(f"   Total Vehicles: {summary['total_vehicles_with_costs']}")
        print(f"   Fleet Capital Cost: £{summary['total_fleet_capital_cost']:,.2f}")
        print(f"   Fleet Operations Cost: £{summary['total_fleet_operations_cost']:,.2f}")
        print(f"   TOTAL FLEET COST: £{summary['total_fleet_cost']:,.2f}")
        print("="*80 + "\n")
        
        return {
            "success": True,
            "summary": summary,
            "vehicles": vehicles_financial,
            "insights": {
                "total_records": len(vehicles_financial),
                "highest_cost_van": vehicles_financial[0] if vehicles_financial else None,
                "cost_structure": "Capital (Lease) + Operations (Service + Maintenance)",
                "money_where": f"{summary['fleet_cost_percentage']['capital']:.1f}% spent on capital costs, {summary['fleet_cost_percentage']['operations']:.1f}% on operations"
            }
        }
        
    except Exception as e:
        print(f"❌ Error generating financial overview: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error generating overview: {str(e)}")
