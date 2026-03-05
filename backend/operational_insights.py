"""
Operational Insights Helper
Combines HSBC Lease data with Service & Maintenance costs
"""

import os
import openpyxl
import csv
from pathlib import Path
from typing import List, Dict

def load_lease_data():
    """Load HSBC lease data from Excel"""
    excel_file = os.path.join(os.path.dirname(__file__), '..', 'HSBC_Leases.xlsx')
    
    if not os.path.exists(excel_file):
        return []
    
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb.active
        
        # Headers in row 2
        headers = [ws.cell(2, i).value for i in range(1, ws.max_column + 1) if ws.cell(2, i).value]
        
        leases = {}
        for row_idx in range(3, ws.max_row + 1):
            row_data = {}
            for col_idx, header in enumerate(headers, 1):
                cell_value = ws.cell(row_idx, col_idx).value
                row_data[header] = cell_value
            
            if row_data.get('Registration Doc '):
                reg = str(row_data.get('Registration Doc ', '')).strip().upper()
                row_data['_trade_group'] = None  # Will be enriched later
                leases[reg] = row_data
        
        return leases
    except Exception as e:
        print(f"❌ Error loading lease data: {e}")
        return {}


def load_service_costs():
    """Load service & maintenance costs from CSV"""
    csv_file = os.path.join(os.path.dirname(__file__), '..', 'frontend', 'public', 'HSBC Leases(Lease Register HSBC).csv')
    
    if not os.path.exists(csv_file):
        return []
    
    try:
        costs = {}
        with open(csv_file, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                registration = (row.get('Registration', '') or '').strip().upper()
                van_number = (row.get('Van #', '') or '').strip()
                
                if not registration:
                    continue
                
                try:
                    total_cost = float((row.get('Total Cost', '') or '0').replace('£', '').replace(',', ''))
                    service_cost = float((row.get('Service', '') or '0').replace('£', '').replace(',', ''))
                    maintenance_cost = float((row.get('Maintenance', '') or '0').replace('£', '').replace(',', ''))
                except:
                    total_cost = service_cost = maintenance_cost = 0
                
                costs[registration] = {
                    'van_number': van_number,
                    'registration': registration,
                    'total_cost': total_cost,
                    'service_cost': service_cost,
                    'maintenance_cost': maintenance_cost,
                }
        
        return costs
    except Exception as e:
        print(f"❌ Error loading service costs: {e}")
        return {}


def get_operational_insights(trade_group_filter=None) -> List[Dict]:
    """
    Combine lease and service costs by registration number
    Returns operational cost breakdown for all vehicles
    """
    leases = load_lease_data()
    costs = load_service_costs()
    
    combined = []
    
    # Match by registration
    for reg, lease in leases.items():
        service_data = costs.get(reg, {
            'van_number': '',
            'total_cost': 0,
            'service_cost': 0,
            'maintenance_cost': 0,
        })
        
        try:
            lease_cost = float((lease.get('Total Repayment ', '') or '0').replace('£', '').replace(',', '')) if lease.get('Total Repayment ') else 0
        except:
            lease_cost = 0
        
        operational = {
            'registration': reg,
            'van_number': service_data.get('van_number', ''),
            'make_model': lease.get('Make & Model ') or '',
            'identifier': lease.get('Identifier ') or '',
            'trade_group': lease.get('_trade_group') or 'Not Assigned',
            'lease_cost': lease_cost,
            'service_cost': service_data.get('service_cost', 0),
            'maintenance_cost': service_data.get('maintenance_cost', 0),
            'total_operational_cost': lease_cost + service_data.get('total_cost', 0),
        }
        
        # Apply trade group filter if provided
        if trade_group_filter and trade_group_filter != 'all':
            if operational['trade_group'] != trade_group_filter:
                continue
        
        combined.append(operational)
    
    # Sort by total operational cost (highest first)
    combined.sort(key=lambda x: x['total_operational_cost'], reverse=True)
    
    return combined


def get_top_10_expensive_vans(trade_group_filter=None) -> List[Dict]:
    """Get top 10 most expensive vans to operate"""
    insights = get_operational_insights(trade_group_filter)
    return insights[:10]


def get_cost_summary(trade_group_filter=None) -> Dict:
    """Get summary statistics for all operational costs"""
    insights = get_operational_insights(trade_group_filter)
    
    if not insights:
        return {
            'total_vans': 0,
            'total_lease_cost': 0,
            'total_service_cost': 0,
            'total_maintenance_cost': 0,
            'total_operational_cost': 0,
            'avg_operational_cost': 0,
        }
    
    return {
        'total_vans': len(insights),
        'total_lease_cost': sum(v['lease_cost'] for v in insights),
        'total_service_cost': sum(v['service_cost'] for v in insights),
        'total_maintenance_cost': sum(v['maintenance_cost'] for v in insights),
        'total_operational_cost': sum(v['total_operational_cost'] for v in insights),
        'avg_operational_cost': sum(v['total_operational_cost'] for v in insights) / len(insights) if insights else 0,
    }
